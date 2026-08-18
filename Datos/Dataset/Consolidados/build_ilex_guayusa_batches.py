# -*- coding: utf-8 -*-
"""
Consolida BATCH 1 (NEG, hoja unica) y BATCH 2 (POS, hojas Pos_1 + Pos_2 + Hoja1,
las 3 se incluyen como pipelines de procesamiento distintos) del articulo Ilex
guayusa (Garzon et al. 2026) en 01_muestras.csv / 02_muestra_actividad.csv /
03_picos.csv. Ver NOTAS_CONSOLIDACION.md para el detalle de cada decision.
"""
import openpyxl, re, csv, os

BASE = r"C:\Users\Dev-Marcos\Desktop\Inteligencia Artificial\Datos"
XLSX_NEG = os.path.join(BASE, "Dataset", "Batch", "BATCH 1- MZmine_to_R_notame_neg.xlsx")
XLSX_POS = os.path.join(BASE, "Dataset", "Batch", "BATCH 2- MZmine_to_R_notame_pos.xlsx")
OUT = os.path.join(BASE, "Dataset", "Consolidados")

CHAKRA = {"Talag": "Chakra A (Talag)", "Alto Pano": "Chakra B (Alto Pano)", "Alto Tena": "Chakra C (Alto Tena)"}
AGE_LABEL = {"Early": "Early (4-6 anios)", "Medium": "Medium (6-8 anios)", "Late": "Late (8-10 anios)"}
LIGHT_LABEL = {"Light": "Light (+, >=300 PPFD)", "Shade": "Shade (-, 0-200 PPFD)"}
TIPO_MAP = {"Sample": "Sample", "QC": "QC", "Sub_Quality_Control": "SubQC", "Blank": "Blank"}

FIXED_META = dict(
    REINO="Plantae", FILO="Tracheophyta", CLASE_TAXONOMICA="Magnoliopsida", ORDEN="Aquifoliales",
    FAMILIA="Aquifoliaceae", GENERO="Ilex", ESPECIE="guayusa", NOMBRE_CIENTIFICO="Ilex guayusa (Loes.)",
    TIPO_PLANTA="Arbol (perennifolio, hasta 30 m)", CICLO_VIDA="Perenne", HABITO_CRECIMIENTO="Erecto",
    TIPO_CULTIVO="Agroforesteria tradicional (chakra), organico, sin fertilizantes/pesticidas sinteticos",
    PARTE_ESTUDIADA="Hoja",
    METODO_EXTRACCION="Extraccion asistida por ultrasonido (sonicacion 20 min, diluyente acuoso)",
    SOLVENTE_EXTRACCION="Agua con 2.0% acido formico y 10% acetonitrilo",
    COLUMNA_CROMATOGRAFICA="ACQUITY UPLC CSH C18 1.7um (3.0x50mm) + VanGuard precolumn",
)
BOTANICAL_FIELDS = ["REINO", "FILO", "CLASE_TAXONOMICA", "ORDEN", "FAMILIA", "GENERO", "ESPECIE",
                     "NOMBRE_CIENTIFICO", "TIPO_PLANTA", "CICLO_VIDA", "HABITO_CRECIMIENTO",
                     "TIPO_CULTIVO", "PARTE_ESTUDIADA"]

COLS_01 = ["ID_MUESTRA", "ID_MUESTRA_FISICA", "ID_NOTAME", "TIPO_MUESTRA", "LOTE", "SECUENCIA_INYECCION",
           "REINO", "FILO", "CLASE_TAXONOMICA", "ORDEN", "FAMILIA", "GENERO", "ESPECIE", "NOMBRE_CIENTIFICO",
           "TIPO_PLANTA", "CICLO_VIDA", "HABITO_CRECIMIENTO", "TIPO_CULTIVO", "PARTE_ESTUDIADA",
           "ORIGEN_GEOGRAFICO", "EDAD_PLANTA", "EXPOSICION_LUZ",
           "METODO_EXTRACCION", "SOLVENTE_EXTRACCION", "MODO_IONIZACION", "COLUMNA_CROMATOGRAFICA"]
COLS_02 = ["ID_MUESTRA_FISICA", "ACTIVIDAD_BIOLOGICA", "VALOR_RESULTADO"]
COLS_03 = ["ID_MUESTRA", "ORIGEN_PROCESAMIENTO", "ID_FEATURE", "RELACION_MASA_CARGA", "TIEMPO_RETENCION_MINUTOS",
           "ALTURA_PICO", "NOMBRE_METABOLITO", "NIVEL_IDENTIFICACION", "CLASE_QUIMICA", "SUPERCLASE_QUIMICA"]

REF_AOX = ("Garzon et al. 2026, J. Chromatogr. A 1769:466732 (revision bibliografica citada en el articulo, "
           "refs. 5,6,8: compuestos fenolicos y CQA de I. guayusa); no medido directamente en este estudio.")
REF_AI = ("Garzon et al. 2026, J. Chromatogr. A 1769:466732 (revision bibliografica citada en el articulo, "
          "ref. 7: isomeros de acido cafeoilquinico); no medido directamente en este estudio.")

suffix_re = re.compile(r"_THOM\d+\.CDF Peak height$")


def clean_id_muestra(raw_header):
    return suffix_re.sub("", raw_header)


def clean_id_fisica(id_muestra):
    return re.sub(r"_(NEG|POS)$", "", id_muestra)


def origen_geo(loc):
    c = CHAKRA.get(loc)
    return f"Napo, Ecuador - {c}" if c else ""


muestras_rows = []
muestras_seen_id = set()
actividad_rows = []
actividad_seen = set()
id_fisica_by_mode = {"NEG": set(), "POS": set()}

picos_f = open(os.path.join(OUT, "03_picos.csv"), "w", newline="", encoding="utf-8")
picos_writer = csv.writer(picos_f)
picos_writer.writerow(COLS_03)
total_picos = 0
resumen = []


def add_muestra_row(id_muestra, id_fisica, id_notame, tipo, lote, secuencia,
                     loc_val, age_val, light_val, modo_ionizacion):
    if id_muestra in muestras_seen_id:
        return
    muestras_seen_id.add(id_muestra)
    row = dict(FIXED_META)
    row.update(
        ID_MUESTRA=id_muestra, ID_MUESTRA_FISICA=id_fisica, ID_NOTAME=id_notame,
        TIPO_MUESTRA=tipo, LOTE=lote, SECUENCIA_INYECCION=secuencia,
        ORIGEN_GEOGRAFICO=origen_geo(loc_val), EDAD_PLANTA=AGE_LABEL.get(age_val, ""),
        EXPOSICION_LUZ=LIGHT_LABEL.get(light_val, ""), MODO_IONIZACION=modo_ionizacion,
    )
    if tipo == "Blank":
        for field in BOTANICAL_FIELDS:
            row[field] = ""
    muestras_rows.append(row)
    if tipo == "Sample" and id_fisica not in actividad_seen:
        actividad_seen.add(id_fisica)
        # Nota: la procedencia de estos 2 valores (literatura general, no medida
        # en este estudio; ver REF_AOX/REF_AI arriba) ya no se guarda como columna
        # -- queda documentada en NOTAS_CONSOLIDACION.md, seccion 5.
        actividad_rows.append(dict(ID_MUESTRA_FISICA=id_fisica, ACTIVIDAD_BIOLOGICA="antioxidante", VALOR_RESULTADO=""))
        actividad_rows.append(dict(ID_MUESTRA_FISICA=id_fisica, ACTIVIDAD_BIOLOGICA="antiinflamatorio", VALOR_RESULTADO=""))


def process_notame_sheet(ws, pipeline_label, modo_ionizacion, lote, sample_col_start,
                          has_identification_level, meta_labels, register_muestras):
    """meta_labels: lista de labels de metadata en el orden de las filas 0..len-1,
    leidos en la columna (sample_col_start - 1)."""
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    n_meta_rows = len(meta_labels)
    meta = {}
    for i, label in enumerate(meta_labels):
        actual_label = rows[i][sample_col_start - 1]
        meta[label] = rows[i][sample_col_start:]
    header_row = rows[n_meta_rows]
    run_headers = [h for h in header_row[sample_col_start:] if h]
    n = len(run_headers)
    id_muestra = [clean_id_muestra(h) for h in run_headers]
    id_fisica = [clean_id_fisica(x) for x in id_muestra]

    if register_muestras:
        id_notame = list(meta.get("Sample_ID", [""] * n))
        tipo = [TIPO_MAP.get(a, a) for a in meta["Analysis_group"][:n]]
        loc = list(meta["Location_Factor"][:n])
        light = list(meta["Light_Factor"][:n])
        age = list(meta["Age_Factor"][:n])
        inj = [str(v).strip() for v in meta["Injection_order"][:n]] if "Injection_order" in meta \
            else [str(v).strip() for v in meta["Injection order"][:n]]
        for i in range(n):
            add_muestra_row(id_muestra[i], id_fisica[i],
                             id_notame[i] if i < len(id_notame) else "",
                             tipo[i], lote, inj[i], loc[i], age[i], light[i], modo_ionizacion)
            id_fisica_by_mode[modo_ionizacion].add(id_fisica[i])

    feature_rows = rows[n_meta_rows + 1:]
    global total_picos
    written = 0
    for fr in feature_rows:
        if fr[0] is None:
            continue
        row_id, mz, rt, metab = fr[0], fr[1], fr[2], fr[3]
        if has_identification_level:
            ident, clase, superclase = fr[4], fr[5], fr[6]
        else:
            ident, clase, superclase = None, fr[4], fr[5]
        heights = fr[sample_col_start:sample_col_start + n]
        for i in range(n):
            picos_writer.writerow([id_muestra[i], pipeline_label, row_id, mz, rt, heights[i],
                                    metab or "", ident or "", clase or "", superclase or ""])
            written += 1
            total_picos += 1
    resumen.append((pipeline_label, n, len(feature_rows), written))


# ---------------- BATCH 1 (NEG, hoja unica) ----------------
wb_neg = openpyxl.load_workbook(XLSX_NEG, read_only=True, data_only=True)
process_notame_sheet(
    wb_neg["Sheet 1"], "BATCH1_NEG", "NEG", "BATCH_1", sample_col_start=9,
    has_identification_level=True,
    meta_labels=["Sample_ID", "Location_Factor", "Light_Factor", "Age_Factor", "Analysis_group",
                 "Factor", "Group", "Visit", "Injection_order", "QC"],
    register_muestras=True,
)

# ---------------- BATCH 2 (POS): Pos_2 = pipeline principal, registra las muestras ----------------
wb_pos = openpyxl.load_workbook(XLSX_POS, read_only=True, data_only=True)
process_notame_sheet(
    wb_pos["Pos_2"], "BATCH2_Pos_2", "POS", "BATCH_2", sample_col_start=9,
    has_identification_level=True,
    meta_labels=["Location_Factor", "Light_Factor", "Age_Factor", "Analysis_group",
                 "Factor", "Group", "Visit", "Injection_order"],
    register_muestras=True,
)

# ---------------- BATCH 2 (POS): Pos_1 y Hoja1 = pipelines alternativos, mismos extractos ----------------
process_notame_sheet(
    wb_pos["Pos_1"], "BATCH2_Pos_1", "POS", "BATCH_2", sample_col_start=8,
    has_identification_level=False,
    meta_labels=["Location_Factor", "Light_Factor", "Age_Factor", "Analysis_group",
                 "Factor", "Group", "Visit", "Injection_order"],
    register_muestras=False,
)
process_notame_sheet(
    wb_pos["Hoja1"], "BATCH2_Hoja1", "POS", "BATCH_2", sample_col_start=9,
    has_identification_level=True,
    meta_labels=["Location_Factor", "Light_Factor", "Age_Factor", "Analysis_group",
                 "Factor", "Group", "Visit", "Injection_order"],
    register_muestras=False,
)

picos_f.close()

# ---------------- Escribir 01 y 02 ----------------
with open(os.path.join(OUT, "01_muestras.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=COLS_01)
    w.writeheader()
    w.writerows(muestras_rows)

with open(os.path.join(OUT, "02_muestra_actividad.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=COLS_02)
    w.writeheader()
    w.writerows(actividad_rows)

only_neg = id_fisica_by_mode["NEG"] - id_fisica_by_mode["POS"]
only_pos = id_fisica_by_mode["POS"] - id_fisica_by_mode["NEG"]
print("Resumen por pipeline (pipeline, n_muestras, n_features, filas_escritas):")
for r in resumen:
    print(" ", r)
print()
print("Total filas 01_muestras.csv:", len(muestras_rows))
print("Total filas 02_muestra_actividad.csv:", len(actividad_rows))
print("Total filas 03_picos.csv:", total_picos)
print("ID_MUESTRA_FISICA solo en NEG (no en POS):", sorted(only_neg))
print("ID_MUESTRA_FISICA solo en POS (no en NEG):", sorted(only_pos))

# -*- coding: utf-8 -*-
"""
Script de construccion de Datos/Dataset/Consolidados a partir de:
  - Datos/Articulos/01_ilex_guayusa.pdf        (ART_01: Garzon et al. 2026, J. Chromatogr. A 1769:466732)
  - Datos/Dataset/Batch/BATCH_01_NEG_ilex_guayusa.xlsx  (hoja 'Sheet 1', formato notame)
  - Datos/Dataset/Batch/BATCH_02_POS_ilex_guayusa.xlsx  (hojas 'Pos_1','Pos_2','Hoja1', formato notame)

Reconstruye TODAS las tablas de Consolidados siguiendo el esquema de Datos/Base/
(mismos nombres de tabla y catalogo, todo ID entero, catalogos agrupados en
subcarpetas). No toca 00_articulos.csv ni 01_batches.csv (ya existian).

Notas de diseno especificas de este dataset real (no estan en el molde generico):
  - SECUENCIA_INYECCION y VISITA se agregan como columnas extra en 03_muestras.csv
    (Injection_order / Visit del Excel) -- no rompen el molde, solo lo extienden.
  - El batch POS trae 3 pipelines de procesamiento (Pos_1, Pos_2, Hoja1) cuyos
    row_ID no son comparables entre si -> catalogo nuevo ID_ORIGEN_PROCESAMIENTO
    en 06_picos.csv (ya reflejado tambien en el molde de Base).
  - El nivel de identificacion (I/II) de los 16 metabolitos nombrados se tomo de
    la Tabla 3 del articulo (fuente mas confiable que el Excel, que lo deja
    vacio o inconsistente en algunas hojas/pipelines).
"""
import csv
import openpyxl

BASE = "."
NEG_XLSX = "../Batch/BATCH_01_NEG_ilex_guayusa.xlsx"
POS_XLSX = "../Batch/BATCH_02_POS_ilex_guayusa.xlsx"


def w(path, header, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(header)
        wr.writerows(rows)
    print(f"  {path}: {len(rows)} filas")


# ---------------------------------------------------------------------------
# 1. Catalogos de especie (estaticos, 1 fila: Ilex guayusa)
# ---------------------------------------------------------------------------
print("=== Catalogos_Especie ===")
w("Catalogos_Especie/reinos.csv", ["ID_REINO", "REINO", "DESCRIPCION"], [[1, "Plantae", ""]])
w("Catalogos_Especie/filos.csv", ["ID_FILO", "FILO", "DESCRIPCION"], [[1, "Tracheophyta", ""]])
w("Catalogos_Especie/clases_taxonomicas.csv", ["ID_CLASE_TAXONOMICA", "CLASE_TAXONOMICA", "DESCRIPCION"], [[1, "Magnoliopsida", ""]])
w("Catalogos_Especie/ordenes.csv", ["ID_ORDEN", "ORDEN", "DESCRIPCION"], [[1, "Aquifoliales", ""]])
w("Catalogos_Especie/familias.csv", ["ID_FAMILIA", "FAMILIA", "DESCRIPCION"], [[1, "Aquifoliaceae", ""]])
w("Catalogos_Especie/generos.csv", ["ID_GENERO", "GENERO", "DESCRIPCION"], [[1, "Ilex", ""]])
w("Catalogos_Especie/tipos_planta.csv", ["ID_TIPO_PLANTA", "TIPO_PLANTA", "DESCRIPCION"], [[1, "Arbol", ""]])
w("Catalogos_Especie/ciclos_vida.csv", ["ID_CICLO_VIDA", "CICLO_VIDA", "DESCRIPCION"], [[1, "Perenne", ""]])
w("Catalogos_Especie/habitos_crecimiento.csv", ["ID_HABITO_CRECIMIENTO", "HABITO_CRECIMIENTO", "DESCRIPCION"], [[1, "Erecta", ""]])

# ---------------------------------------------------------------------------
# 2. 02_especies.csv (hub)
# ---------------------------------------------------------------------------
print("=== 02_especies.csv ===")
w(
    "02_especies.csv",
    ["ID_ESPECIE", "NOMBRE_CIENTIFICO", "ID_REINO", "ID_FILO", "ID_CLASE_TAXONOMICA",
     "ID_ORDEN", "ID_FAMILIA", "ID_GENERO", "ESPECIE", "ID_TIPO_PLANTA", "ID_CICLO_VIDA", "ID_HABITO_CRECIMIENTO"],
    [[1, "Ilex guayusa", 1, 1, 1, 1, 1, 1, "guayusa", 1, 1, 1]],
)

# ---------------------------------------------------------------------------
# 3. Catalogos de muestras (estaticos + factores)
# ---------------------------------------------------------------------------
print("=== Catalogos_Muestras ===")
w("Catalogos_Muestras/tipos_muestra.csv", ["ID_TIPO_MUESTRA", "TIPO_MUESTRA", "DESCRIPCION"], [
    [1, "Sample", "Muestra real de la planta"],
    [2, "Blank", "Blanco de ruido (agua ultrapura), sin especie real"],
    [3, "QC", "Pool de control de calidad, mezcla de todas las muestras"],
    [4, "SubQC", "Pool de control por nivel de un factor (ubicacion, edad o luz)"],
])
w("Catalogos_Muestras/tipos_cultivo.csv", ["ID_TIPO_CULTIVO", "TIPO_CULTIVO", "DESCRIPCION"],
  [[1, "Agroforesteria tradicional organica (chakra)", "Sistema agroforestal indigena, sin fertilizantes ni pesticidas sinteticos"]])
w("Catalogos_Muestras/partes_planta.csv", ["ID_PARTE_PLANTA", "PARTE_PLANTA", "DESCRIPCION"], [[1, "Hoja", ""]])
w("Catalogos_Muestras/ubicaciones.csv", ["ID_UBICACION", "UBICACION", "DESCRIPCION"], [
    [1, "Chakra A (Talag), Napo, Ecuador", ""],
    [2, "Chakra B (Alto Pano), Napo, Ecuador", ""],
    [3, "Chakra C (Alto Tena), Napo, Ecuador", ""],
    [4, "Pool multi-chakra (Talag, Alto Pano, Alto Tena)", "Usado en pools QC/SubQC que mezclan las 3 ubicaciones"],
])
w("Catalogos_Muestras/metodos_extraccion.csv", ["ID_METODO_EXTRACCION", "METODO_EXTRACCION", "DESCRIPCION"],
  [[1, "Ultrasonido (sonicacion 20 min, agitacion manual)", ""]])
w("Catalogos_Muestras/solventes_extraccion.csv", ["ID_SOLVENTE", "SOLVENTE_EXTRACCION", "DESCRIPCION"],
  [[1, "Agua con 2% acido formico y 10% acetonitrilo", ""]])
w("Catalogos_Muestras/columnas_cromatograficas.csv", ["ID_COLUMNA", "COLUMNA_CROMATOGRAFICA", "DESCRIPCION"],
  [[1, "Fase Reversa (RP) - ACQUITY UPLC CSH C18 1.7um (3.0x50mm)", ""]])
w("Catalogos_Muestras/factores_experimentales.csv", ["ID_FACTOR", "FACTOR", "DESCRIPCION"], [
    [1, "Exposicion de luz", "Light_Factor del diseno experimental"],
    [2, "Edad de la planta", "Age_Factor del diseno experimental"],
])

# ---------------------------------------------------------------------------
# 4. Leer metadata de muestras desde los Excel (formato notame)
# ---------------------------------------------------------------------------
print("=== leyendo metadata de muestras ===")


def read_meta_block(ws, header_row_idx, sample_start_col, label_col, n_meta_rows, suffix, sample_end_col=None):
    """Devuelve lista de dicts: 1 por columna de muestra, con las filas de metadata."""
    meta_rows = list(ws.iter_rows(min_row=1, max_row=header_row_idx - 1, values_only=True))
    header_vals = next(r for r in ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    samples = []
    end = sample_end_col if sample_end_col is not None else len(header_vals)
    for c in range(sample_start_col, end):
        code_raw = header_vals[c]
        if not code_raw or suffix not in code_raw:
            continue
        codigo_muestra = code_raw.split(suffix)[0]
        d = {"CODIGO_MUESTRA": codigo_muestra, "_col": c}
        for r in meta_rows:
            label = r[label_col]
            if label is None:
                continue
            d[label] = r[c]
        samples.append(d)
    return samples


wb_neg = openpyxl.load_workbook(NEG_XLSX, read_only=True, data_only=True)
ws_neg = wb_neg["Sheet 1"]
neg_samples = read_meta_block(ws_neg, header_row_idx=11, sample_start_col=9, label_col=8, n_meta_rows=10, suffix="_THOM101.CDF")
for d in neg_samples:
    d["BATCH"] = "NEG"
print(f"  NEG: {len(neg_samples)} muestras")

wb_pos = openpyxl.load_workbook(POS_XLSX, read_only=True, data_only=True)
ws_pos1 = wb_pos["Pos_1"]
pos_samples = read_meta_block(ws_pos1, header_row_idx=9, sample_start_col=8, label_col=7, n_meta_rows=8, suffix="_THOM01.CDF")
for d in pos_samples:
    d["BATCH"] = "POS"
print(f"  POS: {len(pos_samples)} muestras")

all_samples = neg_samples + pos_samples

# ---------------------------------------------------------------------------
# 5. Asignar ID_MUESTRA / ID_MUESTRA_FISICA (enteros) y armar 03_muestras.csv
# ---------------------------------------------------------------------------
print("=== armando 03_muestras.csv y 04_muestra_factor.csv ===")

TM = {"Sample": 1, "Blank": 2, "QC": 3, "Sub_Quality_Control": 4}
UBI = {"Talag": 1, "Alto Pano": 2, "Alto Tena": 3}  # resto (QC/Other/Blank) se resuelve aparte

LIGHT_ES = {"Shade": "Sombra", "Light": "Luz"}
AGE_ES = {"Early": "Temprana (4-6 anios)", "Eary": "Temprana (4-6 anios)",
          "Medium": "Media (6-8 anios)", "Late": "Tardia (8-10 anios)"}

phys_id_map = {}  # codigo_fisico -> ID_MUESTRA_FISICA
next_phys_id = 1
muestras_rows = []
factor_rows = []

for idx, d in enumerate(all_samples, start=1):
    codigo_muestra = d["CODIGO_MUESTRA"]
    codigo_fisico = codigo_muestra.rsplit("_", 1)[0]  # quita _NEG / _POS
    if codigo_fisico not in phys_id_map:
        phys_id_map[codigo_fisico] = next_phys_id
        next_phys_id += 1
    id_muestra_fisica = phys_id_map[codigo_fisico]

    analysis_group = d.get("Analysis_group")
    id_tipo_muestra = TM.get(analysis_group)
    if id_tipo_muestra is None:
        raise ValueError(f"Analysis_group desconocido: {analysis_group!r} en {codigo_muestra}")

    es_blank = id_tipo_muestra == 2
    id_especie = "" if es_blank else 1
    id_tipo_cultivo = "" if es_blank else 1
    id_parte_planta = "" if es_blank else 1
    id_metodo = 1
    id_solvente = 1
    id_columna = 1

    loc = d.get("Location_Factor")
    if es_blank:
        id_ubicacion = ""
    elif loc in UBI:
        id_ubicacion = UBI[loc]
    else:  # 'QC' o 'Other enviroment factors' -> pool mezclado
        id_ubicacion = 4

    lote = "BATCH_01_NEG" if d["BATCH"] == "NEG" else "BATCH_02_POS"
    modo_ion = d["BATCH"]

    visita = d.get("Visit", "")
    inj = d.get("Injection_order", "")
    if isinstance(inj, str):
        inj = inj.strip()

    muestras_rows.append([
        idx, codigo_muestra, id_muestra_fisica, codigo_fisico, id_tipo_muestra, lote,
        id_especie, id_tipo_cultivo, id_parte_planta, id_ubicacion,
        id_metodo, id_solvente, modo_ion, id_columna, visita, inj,
    ])

    # --- factores experimentales: solo si el valor es especifico (no QC/Blank/Other) ---
    light = d.get("Light_Factor")
    if light in LIGHT_ES:
        factor_rows.append([idx, 1, LIGHT_ES[light]])
    age = d.get("Age_Factor")
    if age in AGE_ES:
        factor_rows.append([idx, 2, AGE_ES[age]])

w(
    "03_muestras.csv",
    ["ID_MUESTRA", "CODIGO_MUESTRA", "ID_MUESTRA_FISICA", "CODIGO_MUESTRA_FISICA", "ID_TIPO_MUESTRA", "LOTE",
     "ID_ESPECIE", "ID_TIPO_CULTIVO", "ID_PARTE_PLANTA", "ID_UBICACION",
     "ID_METODO_EXTRACCION", "ID_SOLVENTE", "MODO_IONIZACION", "ID_COLUMNA", "VISITA", "SECUENCIA_INYECCION"],
    muestras_rows,
)
w("04_muestra_factor.csv", ["ID_MUESTRA", "ID_FACTOR", "VALOR"], factor_rows)

# mapa CODIGO_MUESTRA -> ID_MUESTRA, para usar al leer los picos
codigo_a_id_muestra = {r[1]: r[0] for r in muestras_rows}

# ---------------------------------------------------------------------------
# 6. 05_especie_actividad.csv -- vacio (no se dio fuente de actividad biologica esta vez)
# ---------------------------------------------------------------------------
print("=== catalogos de actividad biologica (vacios, sin fuente esta vez) ===")
for fname, idcol, valcol in [
    ("actividades_biologicas.csv", "ID_ACTIVIDAD", "ACTIVIDAD_BIOLOGICA"),
    ("objetivos_actividad.csv", "ID_OBJETIVO", "OBJETIVO"),
    ("metricas_ensayo.csv", "ID_METRICA", "METRICA"),
    ("unidades.csv", "ID_UNIDAD", "UNIDAD"),
    ("condiciones_ensayo.csv", "ID_CONDICION_ENSAYO", "CONDICION_ENSAYO"),
    ("referencias.csv", "ID_REFERENCIA", "REFERENCIA_CITA"),
]:
    w(f"Catalogos_Actividad_Biologica/{fname}", [idcol, valcol, "DESCRIPCION"], [])

w("05_especie_actividad.csv",
  ["ID_ESPECIE", "ID_ACTIVIDAD", "ID_OBJETIVO", "ID_METRICA", "VALOR_NUMERICO", "ID_UNIDAD", "ID_CONDICION_ENSAYO", "ID_REFERENCIA"],
  [])

# ---------------------------------------------------------------------------
# 7. Catalogo de metabolitos (niveles tomados de la Tabla 3 del articulo) y niveles/origenes
# ---------------------------------------------------------------------------
print("=== Catalogos_Picos ===")

# (nombre_excel, clase, superclase, nivel_segun_articulo)
METABOLITOS = [
    ("Acetyl tributyl citrate", "Carboxylic acids and derivatives", "Organic acids and derivatives", "II"),
    ("Caffeine", "Imidazopyrimidines", "Organoheterocyclic compounds", "I"),
    ("Chlorogenic Acid", "Organooxygen compounds", "Organic oxygen compounds", "I"),
    ("Chlorogenic Acid quinone", "Organooxygen compounds", "Organic oxygen compounds", "II"),
    ("Hyperoside", "Flavonoids", "Phenylpropanoids and polyketides", "II"),
    ("Kaempferol-3-O-glucoside", "Flavonoids", "Phenylpropanoids and polyketides", "II"),
    ("Keracyanine", "Flavonoids", "Phenylpropanoids and polyketides", "II"),
    ("Lauryl diethanolamide", "Fatty Acyls", "Lipids and lipid-like molecules", "II"),
    ("Neochlorogenic Acid", "Organooxygen compounds", "Organic oxygen compounds", "I"),
    ("Plantaginin", "Flavonoids", "Phenylpropanoids and polyketides", "II"),
    ("Quinic acid", "Organooxygen compounds", "Organic oxygen compounds", "II"),
    ("Rutin", "Flavonoids", "Phenylpropanoids and polyketides", "I"),
    ("Sucrose", "Organooxygen compounds", "Organic oxygen compounds", "II"),
    ("Theobromine", "Imidazopyrimidines", "Organoheterocyclic compounds", "I"),
    ("Trehalose", "Organooxygen compounds", "Organic oxygen compounds", "II"),
    ("Valerophenone", "Organic oxygen compounds", "Organic compounds", "II"),
]
metab_id = {}
metab_rows = []
for i, (nombre, clase, superclase, _nivel) in enumerate(METABOLITOS, start=1):
    metab_id[nombre] = i
    metab_rows.append([i, nombre, clase, superclase, ""])
w("Catalogos_Picos/metabolitos.csv", ["ID_METABOLITO", "NOMBRE_METABOLITO", "CLASE_QUIMICA", "SUPERCLASE_QUIMICA", "DESCRIPCION"], metab_rows)

nivel_id = {"I": 1, "II": 2, "III": 3, "IV": 4}
w("Catalogos_Picos/niveles_identificacion.csv", ["ID_NIVEL", "NIVEL_IDENTIFICACION", "DESCRIPCION"], [
    [1, "I", "Confirmado con estandar autentico"],
    [2, "II", "Anotado putativamente (libreria espectral)"],
    [3, "III", "Clase quimica putativa"],
    [4, "IV", "Desconocido"],
])
nombre_a_nivel = {nombre: nivel_id[niv] for nombre, _, _, niv in METABOLITOS}

ORIGENES = ["BATCH1_NEG", "BATCH2_Pos_1", "BATCH2_Pos_2", "BATCH2_Hoja1"]
origen_id = {name: i + 1 for i, name in enumerate(ORIGENES)}
w("Catalogos_Picos/origenes_procesamiento.csv", ["ID_ORIGEN_PROCESAMIENTO", "ORIGEN_PROCESAMIENTO", "DESCRIPCION"], [
    [1, "BATCH1_NEG", "Hoja 'Sheet 1' del batch NEG"],
    [2, "BATCH2_Pos_1", "Hoja 'Pos_1' del batch POS (FastDDA 5 iones, merge con full-scan)"],
    [3, "BATCH2_Pos_2", "Hoja 'Pos_2' del batch POS (FastDDA 15 iones, merge con full-scan)"],
    [4, "BATCH2_Hoja1", "Hoja 'Hoja1' del batch POS (pipeline adicional)"],
])

# ---------------------------------------------------------------------------
# 8. 06_picos.csv -- streaming, ~1.28M filas
# ---------------------------------------------------------------------------
print("=== 06_picos.csv (streaming, puede tardar) ===")

PIPELINES = [
    # (workbook, sheet, header_row, cols: dict nombre->indice, id_origen, suffix)
    (wb_neg, "Sheet 1", 11,
     {"id": 0, "mz": 1, "rt": 2, "met": 3, "level": 4, "class": 5, "super": 6},
     origen_id["BATCH1_NEG"], "_THOM101.CDF"),
    (wb_pos, "Pos_1", 9,
     {"id": 0, "mz": 1, "rt": 2, "met": 3, "level": None, "class": 4, "super": 5},
     origen_id["BATCH2_Pos_1"], "_THOM01.CDF"),
    (wb_pos, "Pos_2", 9,
     {"id": 0, "mz": 1, "rt": 2, "met": 3, "level": 4, "class": 5, "super": 6},
     origen_id["BATCH2_Pos_2"], "_THOM01.CDF"),
    (wb_pos, "Hoja1", 9,
     {"id": 0, "mz": 1, "rt": 2, "met": 3, "level": 4, "class": 5, "super": 6},
     origen_id["BATCH2_Hoja1"], "_THOM01.CDF"),
]

total_rows = 0
n_feat_cols = 9  # todas las hojas tienen 8 o 9 columnas de feature antes de las muestras; usamos el max real por hoja abajo

with open("06_picos.csv", "w", encoding="utf-8-sig", newline="") as fout:
    writer = csv.writer(fout)
    writer.writerow(["ID_MUESTRA", "ID_FEATURE", "ID_ORIGEN_PROCESAMIENTO", "RELACION_MASA_CARGA",
                      "TIEMPO_RETENCION_MINUTOS", "ALTURA_PICO", "ID_METABOLITO", "ID_NIVEL"])

    for wb, sheet_name, header_row, cols, id_origen, suffix in PIPELINES:
        ws = wb[sheet_name]
        header_vals = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        n_feat = max(v for v in cols.values() if v is not None) + 1
        # columnas de muestra: las que contienen 'Peak height' y con codigo no vacio
        sample_cols = []  # (col_idx_0based, id_muestra)
        for c in range(n_feat, len(header_vals)):
            hv = header_vals[c]
            if not hv or "Peak height" not in hv or suffix not in hv:
                continue
            codigo = hv.split(suffix)[0]
            id_m = codigo_a_id_muestra.get(codigo)
            if id_m is None:
                raise ValueError(f"Codigo de muestra sin match en {sheet_name}: {codigo!r}")
            sample_cols.append((c, id_m))

        n_rows_sheet = 0
        rows_out = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            feat = row[:n_feat]
            row_id = feat[cols["id"]]
            mz = feat[cols["mz"]]
            rt = feat[cols["rt"]]
            met = feat[cols["met"]] if feat[cols["met"]] else None
            if met:
                id_metab = metab_id.get(met, "")
                id_nivel = nombre_a_nivel.get(met, 4)
            else:
                id_metab = ""
                id_nivel = 4
            for c, id_m in sample_cols:
                altura = row[c]
                rows_out.append([id_m, row_id, id_origen, mz, rt, altura, id_metab, id_nivel])
            n_rows_sheet += 1
            if len(rows_out) >= 20000:
                writer.writerows(rows_out)
                total_rows += len(rows_out)
                rows_out = []
        if rows_out:
            writer.writerows(rows_out)
            total_rows += len(rows_out)
        print(f"  {sheet_name}: {n_rows_sheet} features x {len(sample_cols)} muestras")

print(f"06_picos.csv: {total_rows} filas totales")

wb_neg.close()
wb_pos.close()
print("=== LISTO ===")
